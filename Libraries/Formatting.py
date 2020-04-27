import os
from openpyxl import load_workbook
from Libraries.Styles import *

class Format_TestResultsForFacebookLogin():

    def __init__(self,file):
        self.file = file

    def formatTable(self):
        filePath = os.path.join(os.getcwd(), self.file)
        DataAndResults = load_workbook(filePath)

        Sheet1 = DataAndResults["DataForFacebookLogin"]

        delete = DataAndResults.get_sheet_by_name("TestResultsForFacebookLogin")
        DataAndResults.remove_sheet(delete)

        Sheet2 = DataAndResults.create_sheet("TestResultsForFacebookLogin")

        Sheet2.column_dimensions['A'].width = 30.0
        Sheet2.column_dimensions['B'].width = 30.0
        Sheet2.column_dimensions['C'].width = 20.0
        Sheet2.column_dimensions['D'].width = 20.0

        Sheet2["A1"].alignment = position_left()
        Sheet2["A1"].font = styleBold()
        Sheet2["A1"] = "username"
        Sheet2["B1"].alignment = position_left()
        Sheet2["B1"].font = styleBold()
        Sheet2["B1"] = "password"
        Sheet2["C1"].alignment = position_center()
        Sheet2["C1"].font = styleBold()
        Sheet2["C1"] = "correctness"
        Sheet2["D1"].font = styleBold()
        Sheet2["D1"].alignment = position_center()
        Sheet2["D1"] = "test results"

        number = 2

        for value in Sheet1.iter_rows(min_row=2, min_col=1, max_col=3, values_only=True):
            username = value[0]
            password = value[1]
            correctness = value[2]

            Sheet2["A" + str(number)].alignment = position_left()
            Sheet2["A" + str(number)] = username
            Sheet2["B" + str(number)].alignment = position_left()
            Sheet2["B" + str(number)] = password
            Sheet2["C" + str(number)].alignment = position_center()
            Sheet2["C" + str(number)] = correctness

            number+=1

        DataAndResults.save(self.file)

class Format_other_test_results():

    def __init__(self,file):
        self.file = file

    def formatTable(self):
        filePath = os.path.join(os.getcwd(), self.file)
        DataAndResults = load_workbook(filePath)

        delete = DataAndResults.get_sheet_by_name("other_test_results")
        DataAndResults.remove_sheet(delete)

        Sheet3 = DataAndResults.create_sheet("other_test_results")

        Sheet3.column_dimensions['A'].width = 30.0
        Sheet3.column_dimensions['B'].width = 30.0
        Sheet3.column_dimensions['C'].width = 20.0

        Sheet3["A1"].alignment = position_left()
        Sheet3["A1"].font = styleBold()
        Sheet3["A1"] = "test name"
        Sheet3["B1"].alignment = position_center()
        Sheet3["B1"].font = styleBold()
        Sheet3["B1"] = "class name"
        Sheet3["C1"].alignment = position_center()
        Sheet3["C1"].font = styleBold()
        Sheet3["C1"] = "test results"

        DataAndResults.save(self.file)