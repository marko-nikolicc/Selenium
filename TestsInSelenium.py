import unittest
from selenium import webdriver
from Libraries.Classes import LoginFacebook
import time
from Libraries.Formatting import Format_TestResultsForFacebookLogin,Format_other_test_results
import os
from openpyxl import load_workbook
from Libraries.Styles import *
from Libraries.Functions import WriteResult

class FormatTable():

    formatSheet2 = Format_TestResultsForFacebookLogin("DataAndResults.xlsx")
    formatSheet2.formatTable()
    formatSheet3 = Format_other_test_results("DataAndResults.xlsx")
    formatSheet3.formatTable()

class FacebookLogin(unittest.TestCase):

    def setUp(self):
        self.filePath = os.path.join(os.getcwd(), "DataAndResults.xlsx")
        self.DataAndResults = load_workbook(self.filePath)

    def testloginFacebook(self):
        Sheet1 = self.DataAndResults["DataForFacebookLogin"]
        Sheet2 = self.DataAndResults["TestResultsForFacebookLogin"]
        number = 2

        for value in Sheet1.iter_rows(min_row=2, min_col=1, max_col=3, values_only=True):
            self.driver = webdriver.Chrome()
            self.driver.maximize_window()
            self.driver.get("https://www.facebook.com/")
            self.driver.implicitly_wait(15)
            username = value[0]
            password = value[1]
            correctness = value[2]

            attempt = LoginFacebook(self.driver)
            attempt.setEmail(username)
            time.sleep(1)
            attempt.setPassword(password)
            time.sleep(1)
            attempt.login()
            time.sleep(1)

            if correctness in ["yes","YES","Yes"]:
                try:
                    self.driver.implicitly_wait(15)
                    self.driver.find_element_by_id("userNavigationLabel")
                    Sheet2["D" + str(number)].font = stylePass()
                    Sheet2["D" + str(number)].alignment = position_center()
                    Sheet2["D" + str(number)]="pass"
                except:
                    Sheet2["D" + str(number)].font = styleFail()
                    Sheet2["D" + str(number)].alignment = position_center()
                    Sheet2["D" + str(number)] = "fail"

            elif correctness in ["no","NO","No"]:
                try:
                    self.driver.implicitly_wait(5)
                    self.driver.find_element_by_id("userNavigationLabel")
                    Sheet2["D" + str(number)].font = styleFail()
                    Sheet2["D" + str(number)].alignment = position_center()
                    Sheet2["D" + str(number)] = "fail"
                except:
                    Sheet2["D" + str(number)].font = stylePass()
                    Sheet2["D" + str(number)].alignment = position_center()
                    Sheet2["D" + str(number)] = "pass"

            else:
                Sheet2["D" + str(number)].alignment = position_center()
                Sheet2["D" + str(number)] = "undefined"

            number+=1
            self.driver.close()

    def tearDown(self):
        self.DataAndResults.save("DataAndResults.xlsx")

class NT_Company(unittest.TestCase):

    def setUp(self):
        self.filePath = os.path.join(os.getcwd(), "DataAndResults.xlsx")
        self.DataAndResults = load_workbook(self.filePath)
        self.driver = webdriver.Chrome()
        self.driver.maximize_window()
        self.driver.get("http://www.nt.rs/")
        self.driver.implicitly_wait(15)

    def test_Click_O_NAMA(self):
        try:
            O_NAMA = self.driver.find_element_by_xpath("//a[contains(text(),'O NAMA')]")
            O_NAMA.click()
            assert "NT Company - O KOMPANIJI" in self.driver.title
            result = "pass"
        except AssertionError:
            result = "fail"
        except:
            result = "repeat the test"

        Sheet3 = self.DataAndResults["other_test_results"]
        number = 2
        for value in Sheet3.iter_rows(min_row=1, min_col=1, max_col=3, values_only=True):
            if Sheet3["A" + str(number)].value == "" or Sheet3["A" + str(number)].value == None:
                Sheet3["A" + str(number)] = "test_Click_O_NAMA"
                Sheet3["A" + str(number)].alignment = position_left()
                Sheet3["B" + str(number)] = "NT_Company"
                Sheet3["B" + str(number)].alignment = position_center()
                Sheet3["C" + str(number)] = result
                Sheet3["C" + str(number)].alignment = position_center()
                if result == "pass":
                    Sheet3["C" + str(number)].font = stylePass()
                elif result == "fail":
                    Sheet3["C" + str(number)].font = styleFail()
            else:
                number += 1

    def tearDown(self):
        self.DataAndResults.save("DataAndResults.xlsx")
        self.driver.quit()

if __name__ == "__main__":
    unittest.main()