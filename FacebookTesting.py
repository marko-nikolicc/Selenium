import unittest
from selenium import webdriver
from Libraries.Classes import LoginFacebook
import time
from Libraries.Formatting import Format
import os
from openpyxl import load_workbook
from Libraries.Styles import *


class FacebookLogin(unittest.TestCase):

    def setUp(self):
        formating = Format("DataAndResults.xlsx")
        formating.formatTable()
        self.filePath = os.path.join(os.getcwd(), "DataAndResults.xlsx")
        self.UserPass = load_workbook(self.filePath)

    def testloginFacebook(self):
        Sheet1 = self.UserPass["login"]
        Sheet3 = self.UserPass["test_login_results"]
        number = 2

        for value in Sheet1.iter_rows(min_row=2, min_col=1, max_col=3, values_only=True):
            self.driver = webdriver.Chrome()
            self.driver.maximize_window()
            self.driver.get("https://www.facebook.com/")
            self.driver.implicitly_wait(15)
            username = value[0]
            password = value[1]
            correctness = value[2]

            attempt = LoginFacebook(self.driver)
            attempt.setEmail(username)
            time.sleep(1)
            attempt.setPassword(password)
            time.sleep(1)
            attempt.login()
            time.sleep(1)

            if correctness in ["yes","YES","Yes"]:
                try:
                    self.driver.implicitly_wait(15)
                    home = self.driver.find_element_by_xpath("//a[contains(text(),'Home')]")
                    link = str(home.get_attribute("href"))
                    linkok = "https://www.facebook.com/?ref=tn_tnmn"
                    assert link == linkok
                    Sheet3["D" + str(number)].font = stylePass()
                    Sheet3["D" + str(number)].alignment = position_center()
                    Sheet3["D" + str(number)]="pass"
                except:
                    Sheet3["D" + str(number)].font = styleFail()
                    Sheet3["D" + str(number)].alignment = position_center()
                    Sheet3["D" + str(number)] = "fail"

            elif correctness in ["no","NO","No"]:
                try:
                    self.driver.implicitly_wait(5)
                    home = self.driver.find_element_by_xpath("//a[contains(text(),'Home')]")
                    link = str(home.get_attribute("href"))
                    linkok = "https://www.facebook.com/?ref=tn_tnmn"
                    assert link == linkok
                    Sheet3["D" + str(number)].font = styleFail()
                    Sheet3["D" + str(number)].alignment = position_center()
                    Sheet3["D" + str(number)] = "fail"
                except:
                    Sheet3["D" + str(number)].font = stylePass()
                    Sheet3["D" + str(number)].alignment = position_center()
                    Sheet3["D" + str(number)] = "pass"

            else:
                Sheet3["D" + str(number)].alignment = position_center()
                Sheet3["D" + str(number)] = "undefined"

            number+=1
            self.driver.close()

    def tearDown(self):
        self.UserPass.save("DataAndResults.xlsx")

if __name__ == "__main__":
    unittest.main()